import os
import sys
from openpyxl import load_workbook

# ====== 你的核心逻辑 ======
SLOT_COLS = [4, 7, 8, 9, 11, 13, 14, 17, 21, 22]

ROW_MAP = {
    "Full SKU": 11,
    "Partner SKU": 12,
    "Product Page(s)": 13,
    "Item Name": 14,
    "Category": 15,
    "SubCategory": 16,
    "Current Inventory": 17,
    "QTY Sold": 18,
    "Average First Cost": 19,
    "First Cost Sales": 20,
}


def resource_path(relative_path):
    if hasattr(sys, "_MEIPASS"):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


TEMPLATE_FILE = resource_path("Jeffan - Top 10 - 1.8.26. KB v1.xlsx")


def safe_float(v):
    try:
        return float(v)
    except Exception:
        return float("-inf")


def extract_date_range(ws):
    for row in ws.iter_rows():
        for cell in row:
            val = cell.value
            if isinstance(val, str) and "Date Range" in val:
                return val.strip()
    return "Date Range"


def load_data(raw_file):
    wb = load_workbook(raw_file, data_only=True)

    if "Partner Central" in wb.sheetnames:
        ws = wb["Partner Central"]
    else:
        ws = wb[wb.sheetnames[0]]

    date_range_text = extract_date_range(ws)

    cols = []
    for c in range(4, ws.max_column + 1):
        if ws.cell(10, c).value not in (None, ""):
            cols.append(c)

    items = []
    for c in cols:
        items.append({
            "Full SKU": ws.cell(10, c).value,
            "Partner SKU": ws.cell(11, c).value,
            "Product Page(s)": ws.cell(13, c).value,
            "Item Name": ws.cell(14, c).value,
            "Category": ws.cell(15, c).value,
            "SubCategory": ws.cell(17, c).value,
            "Current Inventory": ws.cell(18, c).value,
            "QTY Sold": ws.cell(19, c).value,
            "Average First Cost": ws.cell(20, c).value,
            "First Cost Sales": ws.cell(21, c).value,
        })

    items.sort(key=lambda x: safe_float(x.get("First Cost Sales")), reverse=True)
    return items[:10], date_range_text


def generate_report(raw_file, output_file):
    items, date_range_text = load_data(raw_file)

    wb = load_workbook(TEMPLATE_FILE)
    ws = wb.active

    # 顶部标题
    ws["A1"] = "Top 10 SKUs"
    ws["A2"] = date_range_text

    # 先清空旧数据
    for col in SLOT_COLS:
        for row in range(11, 21):
            ws.cell(row=row, column=col).value = None

    # 填入新数据
    for i, item in enumerate(items):
        if i >= len(SLOT_COLS):
            break
        col = SLOT_COLS[i]
        for key, row in ROW_MAP.items():
            ws.cell(row=row, column=col).value = item.get(key)

    wb.save(output_file)


# ====== GUI ======
class App(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("GG")
        self.setGeometry(100, 100, 400, 200)

        layout = QVBoxLayout()

        self.label = QLabel("No file selected")
        layout.addWidget(self.label)

        self.select_btn = QPushButton("Select Partner Central File")
        self.select_btn.clicked.connect(self.select_file)
        layout.addWidget(self.select_btn)

        self.generate_btn = QPushButton("Generate Report")
        self.generate_btn.clicked.connect(self.generate)
        layout.addWidget(self.generate_btn)

        self.setLayout(layout)

        self.file_path = None

    def select_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "Select File", "", "Excel Files (*.xlsx)")
        if file:
            self.file_path = file
            self.label.setText(file)

    def generate(self):
        if not self.file_path:
            QMessageBox.warning(self, "Error", "Please select a file first")
            return

        save_path, _ = QFileDialog.getSaveFileName(self, "Save File", "Output.xlsx", "Excel Files (*.xlsx)")
        if not save_path:
            return

        try:
            generate_report(self.file_path, save_path)
            QMessageBox.information(self, "Success", "Report generated!")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = App()
    window.show()
    sys.exit(app.exec())